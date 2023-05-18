import xml.etree.ElementTree as ET
import openpyxl as xl

tree = ET.parse('go_obo.xml')
root = tree.getroot()
go_terms = []


for term in root.iter('term'):
    # Extract the GO id, term name, and definition string
    go_id = term.find('id').text
    term_name = term.find('name').text
    definition = term.find('def/defstr').text
    
    # Check if the definition contains 'autophagosome'
    if definition and 'autophagosome' in definition:
        # Count the number of child nodes
        child_nodes = sum(1 for _ in term.iter('is_a'))
        
        # Add the GO term to the list
        go_terms.append((go_id, term_name, definition, child_nodes))


workbook = xl.Workbook()
sheet = workbook.active

sheet['A1'] = 'GO ID'
sheet['B1'] = 'Term Name'
sheet['C1'] = 'Definition'
sheet['D1'] = 'Number of Child Nodes'


for row, go_term in enumerate(go_terms, start=2):
    sheet.cell(row=row, column=1).value = go_term[0]
    sheet.cell(row=row, column=2).value = go_term[1]
    sheet.cell(row=row, column=3).value = go_term[2]
    sheet.cell(row=row, column=4).value = go_term[3]


workbook.save('autophagosome.xlsx')
